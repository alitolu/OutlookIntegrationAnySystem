from typing import Any, Dict, List
import pandas as pd
import os
from datetime import datetime
import numpy as np
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import re
from utils.fuzzy_matcher import FuzzyMatcher
import pyodbc
from .data_source import DataSource

class ExcelHelper(DataSource):
    def __init__(self, file_path: str, config: Dict):
        self.file_path = file_path
        self.config = config
        self.data = None
        try:
            if os.path.exists(file_path):
                self.df = pd.read_excel(file_path)
            else:
                print(f"Excel dosyası bulunamadı: {file_path}")
                self.df = pd.DataFrame()
        except Exception as e:
            print(f"Excel yükleme hatası: {str(e)}")
            self.df = pd.DataFrame()

    def load_data(self) -> pd.DataFrame:
        try:
            if not os.path.exists(self.file_path):
                print(f"Excel dosyası bulunamadı: {self.file_path}")
                return pd.DataFrame()
            
            self.data = pd.read_excel(self.data_source)
            return self.data
        except Exception as e:
            print(f"Excel yükleme hatası: {str(e)}")
            return pd.DataFrame()

    def save_data(self, df, create_backup=True):
        try:
            if create_backup:
                backup_path = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                df.to_excel(backup_path, index=False)
                
            df.to_excel(self.data_source, index=False)
            return True
        except Exception as e:
            print(f"Excel kaydetme hatası: {str(e)}")
            return False

    def find_awb(self, search_text: str) -> dict:
        try:
            print(f"Arama başladı Excel: {search_text}")
            
            # 1. Config var mı kontrolü
            if not self.config:
                print("Uyarı: Config bulunamadı!")
                return {}

            # 2. Data var mı kontrolü 
            if self.df is None or self.df.empty:
                print("Uyarı: Excel verisi boş!")
                return {}

            # 3. Aranabilir kolonu belirle
            search_column = None
            for col_name, details in self.config.get("datasource", {}).get("column_mappings", {}).items():
                if details.get("searchable", False):
                    search_column = col_name
                    break

            # 4. Arama kolonu yoksa varsayılan kolonu al
            if not search_column:
                search_column = self.config.get("datasource", {}).get("search_column")
                
            # 5. Kolon kontrolü ve arama
            if search_column and search_column in self.df.columns:
                clean_awb = str(search_text).strip().upper()
                print(f"Arama yapılıyor: '{clean_awb}' -> Kolon: '{search_column}'")
                
                # 6. DataFrame'de arama yap
                matches = self.df[self.df[search_column].str.upper().str.contains(
                    clean_awb, 
                    na=False
                )]

                if not matches.empty:
                    # 7. İlk eşleşmeyi al
                    result = matches.iloc[0].to_dict()
                    
                    # 8. Görünür kolonları belirle
                    visible_columns = {
                        col: details.get("display_name")
                        for col, details in self.config.get("datasource", {})
                                                     .get("column_mappings", {}).items()
                        if details.get("visible", True)
                    }
                    
                    # 9. Sonucu formatla
                    formatted_result = {
                        details: result.get(col, '')
                        for col, details in visible_columns.items()
                    }
                    
                 
                    return formatted_result
                else:
                    print(f"'{clean_awb}' için eşleşme bulunamadı")
                    
            else:
                print(f"Uyarı: '{search_column}' sütunu bulunamadı!")
                
        except Exception as e:
            print(f"Excel arama hatası: {str(e)}")
            import traceback
            print(traceback.format_exc())
            
        return {}

    def export_awb_results(self, results, file_name=None):
        """AWB sonuçlarını Excel'e aktar"""
        try:
            # Yeni Excel dosyası adı
            if file_name is None:
                file_name = f"awb_sonuclari_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Config'den görünür kolonları al
            visible_columns = {
                col: details.get("display_name")
                for col, details in self.config.get("datasource", {})
                                             .get("column_mappings", {}).items()
                if details.get("visible", True)
            }

            # DataFrame için veri hazırla
            export_data = []
            for result in results:
                # Excel'den eşleşme verisini al
                excel_data = self.find_awb(result['awb'])
                
                row_data = {
                    details: excel_data.get(col, '')
                    for col, details in visible_columns.items()
                }
                
                # Mail bilgilerini ekle
                row_data.update({
                    'Mail Tarihi': datetime.fromisoformat(result['date']).strftime('%Y-%m-%d %H:%M'),
                    'Mail Konusu': result['subject'],
                    'Eşleşme': result['awb'],
                    'Bulunan Metin': result['matched_text'],
                    'Bulunduğu Yer': result.get('location', 'Mail İçeriği'),
                    'İşlem Tarihi': datetime.now().strftime('%Y-%m-%d %H:%M')
                })
                
                export_data.append(row_data)

            # DataFrame oluştur
            df = pd.DataFrame(export_data)

            # Excel'e kaydet
            df.to_excel(file_name, index=False, sheet_name='Eşleşme Sonuçları')
            
            # Kolon genişliklerini ayarla
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
                for column in df:
                    column_width = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    writer.sheets['Eşleşme Sonuçları'].column_dimensions[chr(65 + col_idx)].width = column_width + 2

            return True, file_name

        except Exception as e:
            print(f"Excel export hatası: {str(e)}")
            return False, str(e)

    def format_excel(self, writer):
        """Excel dosyasını formatla"""
        workbook = writer.book
        worksheet = writer.sheets['AWB Sonuçları']
        
        # Sütun genişliklerini ayarla
        for idx, col in enumerate(self.data.columns):
            max_length = max(
                self.data[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 2
            
        # Başlık formatı
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Koşullu formatlama
        status_colors = {
            'Tamamlandı': '00B050',  # Yeşil
            'Beklemede': 'FFC000',   # Sarı
            'Gecikme': 'FF0000'      # Kırmızı
        }
        
        return writer

    def open_file(self):
        """Excel dosyasını varsayılan uygulamada aç"""
        if os.path.exists(self.filename):
            os.startfile(self.filename)
        else:
            raise FileNotFoundError(f"Excel dosyası bulunamadı: {self.filename}")

    def connect(self) -> bool:
        try:
            self.data = self.load_data()
            return True
        except:
            return False

    def disconnect(self) -> None:
        self.data = None

    def test_connection(self) -> bool:
        return os.path.exists(self.data_source)

    def get_all_data(self) -> List[Dict[str, Any]]:
        """Tüm veriyi liste olarak döndür"""
        try:
            if self.data is None or self.data.empty:
                self.data = self.load_data()
                
            if self.data is not None and not self.data.empty:
                # DataFrame'i dictionary listesine çevir
                records = self.data.to_dict('records')
                
                # Config'den görünür kolonları al
                visible_columns = {
                    col: details.get("display_name")
                    for col, details in self.config.get("datasource", {})
                                                 .get("column_mappings", {}).items()
                    if details.get("visible", True)
                }
                
                # Sadece görünür kolonları içeren dictionary listesi oluştur
                result = []
                for record in records:
                    row_data = {
                        details: record.get(col, '')
                        for col, details in visible_columns.items()
                    }
                    result.append(row_data)
                    
                return result
                
            return []
            
        except Exception as e:
            print(f"Veri alma hatası: {str(e)}")
            return []